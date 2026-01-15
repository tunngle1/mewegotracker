"""
Script to import habits from Excel export.
Run: python import_habits.py
"""
import asyncio
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from sqlalchemy import select

from bot.database import async_session, init_db
from bot.models import User, Habit, ScheduleType


MSK = ZoneInfo("Europe/Moscow")


def parse_datetime(dt_str):
    """Parse datetime string from export."""
    if not dt_str or dt_str == "":
        return None
    try:
        msk_dt = datetime.strptime(str(dt_str), "%Y-%m-%d %H:%M")
        msk_dt = msk_dt.replace(tzinfo=MSK)
        return msk_dt.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
    except Exception:
        return None


def safe_int(value, default=None):
    """Safely convert to int."""
    if value is None or value == "" or value == "None":
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


async def import_habits():
    """Import habits from Excel file."""
    await init_db()
    
    # Load Excel file
    wb = load_workbook("database old/mewego_habits_20260115_1129.xlsx")
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers)}
    print(f"Headers: {headers}")
    
    # Helper to get column value
    def get_col(row, name, default=None):
        idx = col_map.get(name)
        if idx is not None and idx < len(row):
            val = row[idx]
            return val if val and val != "None" else default
        return default
    
    imported = 0
    skipped = 0
    errors = 0
    
    async with async_session() as session:
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                telegram_id = safe_int(get_col(row, "Telegram ID"))
                if not telegram_id:
                    continue
                
                # Find user by Telegram ID
                result = await session.execute(
                    select(User).where(User.telegram_id == telegram_id)
                )
                user = result.scalar_one_or_none()
                
                if not user:
                    print(f"User not found for Telegram ID: {telegram_id}")
                    errors += 1
                    continue
                
                habit_name = get_col(row, "Название привычки")
                if not habit_name:
                    continue
                
                # Check if habit already exists for this user
                result = await session.execute(
                    select(Habit).where(
                        Habit.user_id == user.id,
                        Habit.name == habit_name
                    )
                )
                existing = result.scalar_one_or_none()
                
                if existing:
                    print(f"Skipping existing habit: {habit_name} for user {user.name}")
                    skipped += 1
                    continue
                
                # Parse schedule type
                type_str = get_col(row, "Тип", "Ежедневно")
                if "Ежедневно" in str(type_str):
                    schedule_type = ScheduleType.DAILY
                else:
                    schedule_type = ScheduleType.WEEKLY
                
                # Create habit
                habit = Habit(
                    user_id=user.id,
                    name=habit_name,
                    schedule_type=schedule_type,
                    weekly_target=safe_int(get_col(row, "Цель (раз/нед)"), 7),
                    is_active=(get_col(row, "Активна") == "Да"),
                    created_at=parse_datetime(get_col(row, "Дата создания (МСК)")),
                )
                
                session.add(habit)
                imported += 1
                print(f"Imported habit: {habit_name} for {user.name}")
                
            except Exception as e:
                print(f"Error on row {row_num}: {e}")
                errors += 1
                continue
        
        await session.commit()
    
    print(f"\n=== Habits Import Complete ===")
    print(f"Imported: {imported}")
    print(f"Skipped (already exist): {skipped}")
    print(f"Errors: {errors}")


if __name__ == "__main__":
    asyncio.run(import_habits())
