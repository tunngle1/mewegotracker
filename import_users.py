"""
Script to import users from Excel export back to database.
Run once to restore users: python import_users.py
"""
import asyncio
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from sqlalchemy import select

# Import from bot modules
from bot.database import async_session, init_db
from bot.models import User


MSK = ZoneInfo("Europe/Moscow")


def parse_datetime(dt_str: str):
    """Parse datetime string from export (Moscow time)."""
    if not dt_str or dt_str == "":
        return None
    try:
        # Format: "2026-01-15 11:22"
        msk_dt = datetime.strptime(str(dt_str), "%Y-%m-%d %H:%M")
        msk_dt = msk_dt.replace(tzinfo=MSK)
        # Convert to UTC for storage
        return msk_dt.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
    except Exception:
        return None


def safe_int(value, default=None):
    """Safely convert to int."""
    if value is None or value == "" or value == "None":
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


async def import_users():
    """Import users from Excel file."""
    # Initialize database (creates tables if needed)
    await init_db()
    
    # Load Excel file
    wb = load_workbook("database old/mewego_users_20260115_1122.xlsx")
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    print(f"Total columns: {len(headers)}")
    
    # Build column index map
    col_map = {h: i for i, h in enumerate(headers)}
    print(f"Column map: {col_map}")
    
    imported = 0
    skipped = 0
    errors = 0
    
    async with async_session() as session:
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                telegram_id_val = row[col_map.get("Telegram ID", 1)]
                if not telegram_id_val:
                    continue
                
                telegram_id = int(telegram_id_val)
                
                # Check if user already exists
                result = await session.execute(
                    select(User).where(User.telegram_id == telegram_id)
                )
                existing = result.scalar_one_or_none()
                
                if existing:
                    print(f"Skipping existing user: {telegram_id}")
                    skipped += 1
                    continue
                
                # Get values by column name (safe)
                def get_col(name, default=None):
                    idx = col_map.get(name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return val if val and val != "None" else default
                    return default
                
                # Create user
                user = User(
                    telegram_id=telegram_id,
                    username=get_col("Username"),
                    name=get_col("Имя"),
                    age=safe_int(get_col("Возраст")),
                    city=get_col("Город"),
                    activity_level=get_col("Активность"),
                    goal=get_col("Цель"),
                    current_habit=get_col("Привычка"),
                    custom_habit=get_col("Своя привычка"),
                    day_cycle=safe_int(get_col("День цикла"), 1),
                    reminder_time=get_col("Время напоминания"),
                    onboarding_completed=(get_col("Онбординг") == "Да"),
                    self_identification=get_col("Самоопознавание"),
                    created_at=parse_datetime(get_col("Дата регистрации (МСК)")),
                    last_check_in=parse_datetime(get_col("Последняя отметка (МСК)")),
                )
                
                session.add(user)
                imported += 1
                print(f"Imported: {user.name} ({telegram_id})")
                
            except Exception as e:
                print(f"Error on row {row_num}: {e}")
                errors += 1
                continue
        
        await session.commit()
    
    print(f"\n=== Import Complete ===")
    print(f"Imported: {imported}")
    print(f"Skipped (already exist): {skipped}")
    print(f"Errors: {errors}")


if __name__ == "__main__":
    asyncio.run(import_users())
