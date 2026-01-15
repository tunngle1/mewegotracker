"""Admin panel handlers."""
import csv
import io
import json
import asyncio
from datetime import datetime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, ConversationHandler
from sqlalchemy import select, func

from bot.database import async_session
from bot.models import User, HabitLog, Poll, PollVote
from bot.config import ADMIN_USERNAMES
from bot.messages import HABITS

# States for broadcast conversation
BROADCAST_WAITING_MESSAGE = 100
BROADCAST_CONFIRM = 101

# States for poll creation
POLL_WAITING_QUESTION = 110
POLL_WAITING_OPTIONS = 111
POLL_CONFIRM = 112


def get_admin_panel_keyboard() -> InlineKeyboardMarkup:
    """Клавиатура админ-панели с кнопками команд."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Статистика", callback_data="admin:stats")],
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin:users")],
        [InlineKeyboardButton("📢 Рассылка", callback_data="admin:broadcast")],
        [InlineKeyboardButton("🗳 Создать опрос", callback_data="admin:poll")],
        [InlineKeyboardButton("📊 Мои опросы", callback_data="admin:polls_list")],
        [InlineKeyboardButton("📁 Экспорт пользователей", callback_data="admin:export")],
        [InlineKeyboardButton("📋 Экспорт привычек", callback_data="admin:export_habits")],
    ])


def is_admin(username: str) -> bool:
    """Check if user is admin by username."""
    if not username or not ADMIN_USERNAMES:
        return False
    return username.lower() in ADMIN_USERNAMES


async def admin_stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show admin panel with buttons. Command: /admin or button 🔐 Админ-панель"""
    # Support both message and callback query
    if update.callback_query:
        user = update.callback_query.from_user
        reply_func = update.callback_query.message.reply_text
        await update.callback_query.answer()
    else:
        user = update.effective_user
        reply_func = update.message.reply_text
    
    if not is_admin(user.username):
        if update.callback_query:
            await update.callback_query.answer("❌ Нет доступа", show_alert=True)
        else:
            await reply_func("❌ Нет доступа")
        return
    
    await reply_func(
        "🔐 <b>Админ-панель MeWeGo</b>\n\n"
        "Выбери действие:",
        parse_mode="HTML",
        reply_markup=get_admin_panel_keyboard()
    )


async def admin_callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle admin panel button clicks."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return
    
    action = query.data.split(":")[1]
    
    if action == "stats":
        await show_stats(query)
    elif action == "users":
        await show_users(query)
    elif action == "export":
        await do_export_users(query)
    elif action == "export_habits":
        await do_export_habits(query)
    elif action == "broadcast":
        return await start_broadcast(update, context)


async def show_stats(query) -> None:
    """Show bot statistics."""
    async with async_session() as session:
        # Total users
        total_users = await session.execute(select(func.count(User.id)))
        total_users = total_users.scalar()
        
        # Completed onboarding
        completed = await session.execute(
            select(func.count(User.id)).where(User.onboarding_completed == True)
        )
        completed = completed.scalar()
        
        # Total check-ins
        total_checkins = await session.execute(select(func.count(HabitLog.id)))
        total_checkins = total_checkins.scalar()
        
        # Today's check-ins
        today = datetime.utcnow().date()
        today_checkins = await session.execute(
            select(func.count(HabitLog.id)).where(
                func.date(HabitLog.completed_at) == today
            )
        )
        today_checkins = today_checkins.scalar()
    
    stats_text = (
        "📊 <b>Статистика MeWeGo</b>\n\n"
        f"👥 Всего пользователей: {total_users}\n"
        f"✅ Прошли онбординг: {completed}\n"
        f"📝 Всего отметок: {total_checkins}\n"
        f"📅 Отметок сегодня: {today_checkins}\n"
    )
    
    await query.message.reply_text(
        stats_text, 
        parse_mode="HTML",
        reply_markup=get_admin_panel_keyboard()
    )


async def show_users(query) -> None:
    """Show users list."""
    async with async_session() as session:
        result = await session.execute(
            select(User).order_by(User.created_at.desc()).limit(50)
        )
        users = result.scalars().all()
    
    if not users:
        await query.message.reply_text("Пользователей пока нет")
        return
    
    text = "👥 <b>Последние пользователи:</b>\n\n"
    
    for user in users:
        status = "✅" if user.onboarding_completed else "⏳"
        habit = user.custom_habit if user.current_habit == "custom" else dict(HABITS).get(user.current_habit, "-")
        text += (
            f"{status} <b>{user.name or 'Без имени'}</b>\n"
            f"   📍 {user.city or '-'} | 🎯 {user.goal or '-'}\n"
            f"   🔄 День {user.day_cycle}/30 | ⏰ {user.reminder_time or '-'}\n\n"
        )
    
    await query.message.reply_text(
        text, 
        parse_mode="HTML",
        reply_markup=get_admin_panel_keyboard()
    )


async def users_list_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """List all users. Command: /users"""
    if not is_admin(update.effective_user.username):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    async with async_session() as session:
        result = await session.execute(
            select(User).order_by(User.created_at.desc()).limit(50)
        )
        users = result.scalars().all()
    
    if not users:
        await update.message.reply_text("Пользователей пока нет")
        return
    
    text = "👥 <b>Последние пользователи:</b>\n\n"
    
    for user in users:
        status = "✅" if user.onboarding_completed else "⏳"
        habit = user.custom_habit if user.current_habit == "custom" else dict(HABITS).get(user.current_habit, "-")
        text += (
            f"{status} <b>{user.name or 'Без имени'}</b>\n"
            f"   📍 {user.city or '-'} | 🎯 {user.goal or '-'}\n"
            f"   🔄 День {user.day_cycle}/30 | ⏰ {user.reminder_time or '-'}\n\n"
        )
    
    await update.message.reply_text(text, parse_mode="HTML")


async def do_export_users(query) -> None:
    """Export users via callback button."""
    async with async_session() as session:
        result = await session.execute(select(User))
        users = result.scalars().all()
    
    if not users:
        await query.message.reply_text("Нет данных для экспорта")
        return
    
    from zoneinfo import ZoneInfo
    from datetime import timezone
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    MSK = ZoneInfo("Europe/Moscow")
    
    def to_msk(dt):
        if not dt:
            return ""
        utc_dt = dt.replace(tzinfo=timezone.utc)
        msk_dt = utc_dt.astimezone(MSK)
        return msk_dt.strftime("%Y-%m-%d %H:%M")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = [
        "ID", "Telegram ID", "Username", "Имя", "Возраст", "Город",
        "Активность", "Цель", "Формат тренировок", "Привычка", "Своя привычка",
        "День цикла", "Время напоминания", "Онбординг",
        "Самоопознавание", "Дата регистрации (МСК)", "Последняя отметка (МСК)"
    ]
    ws.append(headers)
    
    for user in users:
        habit_display = dict(HABITS).get(user.current_habit, user.current_habit)
        ws.append([
            user.id,
            user.telegram_id,
            user.username or "",
            user.name or "",
            user.age or "",
            user.city or "",
            user.activity_level or "",
            user.goal or "",
            user.training_preference or "",
            habit_display or "",
            user.custom_habit or "",
            user.day_cycle,
            user.reminder_time or "",
            "Да" if user.onboarding_completed else "Нет",
            user.self_identification or "",
            to_msk(user.created_at),
            to_msk(user.last_check_in)
        ])
    
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"mewego_users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    await query.message.reply_document(
        document=output,
        filename=filename,
        caption=f"📊 Экспорт {len(users)} пользователей"
    )


async def do_export_habits(query) -> None:
    """Export habits via callback button."""
    from sqlalchemy.orm import selectinload
    from bot.models import Habit, ScheduleType, LogStatus
    
    async with async_session() as session:
        result = await session.execute(
            select(User).options(
                selectinload(User.habits).selectinload(Habit.logs)
            )
        )
        users = result.scalars().all()
    
    if not users:
        await query.message.reply_text("Нет данных для экспорта")
        return
    
    from zoneinfo import ZoneInfo
    from datetime import timezone
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    MSK = ZoneInfo("Europe/Moscow")
    
    def to_msk(dt):
        if not dt:
            return ""
        utc_dt = dt.replace(tzinfo=timezone.utc)
        msk_dt = utc_dt.astimezone(MSK)
        return msk_dt.strftime("%Y-%m-%d %H:%M")
    
    wb = Workbook()
    ws_habits = wb.active
    ws_habits.title = "Привычки"
    
    habits_headers = [
        "ID привычки", "ID пользователя", "Telegram ID", "Имя пользователя",
        "Название привычки", "Тип", "Цель (раз/нед)", "Активна",
        "Всего выполнений", "Дата создания (МСК)"
    ]
    ws_habits.append(habits_headers)
    
    total_habits = 0
    for user in users:
        if not user.habits:
            continue
        for habit in user.habits:
            total_habits += 1
            done_count = sum(1 for log in habit.logs if log.status == LogStatus.DONE) if habit.logs else 0
            schedule_type = "Ежедневно" if habit.schedule_type == ScheduleType.DAILY else f"{habit.weekly_target}x в неделю"
            
            ws_habits.append([
                habit.id, user.id, user.telegram_id, user.name or user.username or "",
                habit.name, schedule_type, habit.weekly_target,
                "Да" if habit.is_active else "Нет", done_count, to_msk(habit.created_at),
            ])
    
    ws_logs = wb.create_sheet("Логи")
    logs_headers = [
        "ID лога", "ID привычки", "Название привычки", "ID пользователя",
        "Имя пользователя", "Дата", "Статус", "День цикла", "Время отметки (МСК)"
    ]
    ws_logs.append(logs_headers)
    
    total_logs = 0
    for user in users:
        if not user.habits:
            continue
        for habit in user.habits:
            if not habit.logs:
                continue
            for log in habit.logs:
                total_logs += 1
                status_text = {
                    LogStatus.DONE: "Выполнено",
                    LogStatus.NOT_DONE: "Не сделал",
                    LogStatus.SKIPPED: "Пропуск",
                }.get(log.status, str(log.status))
                
                ws_logs.append([
                    log.id, habit.id, habit.name, user.id, user.name or user.username or "",
                    log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
                    status_text, log.day_cycle or "", to_msk(log.completed_at),
                ])
    
    for ws in [ws_habits, ws_logs]:
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"mewego_habits_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    await query.message.reply_document(
        document=output,
        filename=filename,
        caption=f"📊 Экспорт: {total_habits} привычек, {total_logs} логов"
    )


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Export all users to Excel. Command: /export"""
    if not is_admin(update.effective_user.username):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    async with async_session() as session:
        result = await session.execute(select(User))
        users = result.scalars().all()
    
    if not users:
        await update.message.reply_text("Нет данных для экспорта")
        return
    
    # Moscow timezone
    from zoneinfo import ZoneInfo
    from datetime import timezone
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    MSK = ZoneInfo("Europe/Moscow")
    
    def to_msk(dt):
        """Convert UTC datetime to Moscow time string."""
        if not dt:
            return ""
        utc_dt = dt.replace(tzinfo=timezone.utc)
        msk_dt = utc_dt.astimezone(MSK)
        return msk_dt.strftime("%Y-%m-%d %H:%M")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Header
    headers = [
        "ID", "Telegram ID", "Username", "Имя", "Возраст", "Город",
        "Активность", "Цель", "Формат тренировок", "Привычка", "Своя привычка",
        "День цикла", "Время напоминания", "Онбординг",
        "Самоопознавание", "Дата регистрации (МСК)", "Последняя отметка (МСК)"
    ]
    ws.append(headers)
    
    # Data
    for user in users:
        habit_display = dict(HABITS).get(user.current_habit, user.current_habit)
        ws.append([
            user.id,
            user.telegram_id,
            user.username or "",
            user.name or "",
            user.age or "",
            user.city or "",
            user.activity_level or "",
            user.goal or "",
            user.training_preference or "",
            habit_display or "",
            user.custom_habit or "",
            user.day_cycle,
            user.reminder_time or "",
            "Да" if user.onboarding_completed else "Нет",
            user.self_identification or "",
            to_msk(user.created_at),
            to_msk(user.last_check_in)
        ])
    
    # Auto-size columns
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"mewego_users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    await update.message.reply_document(
        document=output,
        filename=filename,
        caption=f"📊 Экспорт {len(users)} пользователей"
    )


async def export_habits_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Export all habits and logs to Excel. Command: /export_habits"""
    if not is_admin(update.effective_user.username):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    from sqlalchemy.orm import selectinload
    from bot.models import Habit, ScheduleType, LogStatus
    
    async with async_session() as session:
        result = await session.execute(
            select(User)
            .options(
                selectinload(User.habits).selectinload(Habit.logs)
            )
        )
        users = result.scalars().all()
    
    if not users:
        await update.message.reply_text("Нет данных для экспорта")
        return
    
    # Moscow timezone
    from zoneinfo import ZoneInfo
    from datetime import timezone
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    MSK = ZoneInfo("Europe/Moscow")
    
    def to_msk(dt):
        """Convert UTC datetime to Moscow time string."""
        if not dt:
            return ""
        utc_dt = dt.replace(tzinfo=timezone.utc)
        msk_dt = utc_dt.astimezone(MSK)
        return msk_dt.strftime("%Y-%m-%d %H:%M")
    
    # Create Excel workbook
    wb = Workbook()
    
    # ========== Sheet 1: Habits ==========
    ws_habits = wb.active
    ws_habits.title = "Привычки"
    
    habits_headers = [
        "ID привычки", "ID пользователя", "Telegram ID", "Имя пользователя",
        "Название привычки", "Тип", "Цель (раз/нед)", "Активна",
        "Всего выполнений", "Дата создания (МСК)"
    ]
    ws_habits.append(habits_headers)
    
    total_habits = 0
    for user in users:
        if not user.habits:
            continue
        for habit in user.habits:
            total_habits += 1
            done_count = sum(1 for log in habit.logs if log.status == LogStatus.DONE) if habit.logs else 0
            schedule_type = "Ежедневно" if habit.schedule_type == ScheduleType.DAILY else f"{habit.weekly_target}x в неделю"
            
            ws_habits.append([
                habit.id,
                user.id,
                user.telegram_id,
                user.name or user.username or "",
                habit.name,
                schedule_type,
                habit.weekly_target,
                "Да" if habit.is_active else "Нет",
                done_count,
                to_msk(habit.created_at),
            ])
    
    # ========== Sheet 2: Logs ==========
    ws_logs = wb.create_sheet("Логи")
    
    logs_headers = [
        "ID лога", "ID привычки", "Название привычки", "ID пользователя",
        "Имя пользователя", "Дата", "Статус", "День цикла", "Время отметки (МСК)"
    ]
    ws_logs.append(logs_headers)
    
    total_logs = 0
    for user in users:
        if not user.habits:
            continue
        for habit in user.habits:
            if not habit.logs:
                continue
            for log in habit.logs:
                total_logs += 1
                status_text = {
                    LogStatus.DONE: "Выполнено",
                    LogStatus.NOT_DONE: "Не сделал",
                    LogStatus.SKIPPED: "Пропуск",
                }.get(log.status, str(log.status))
                
                ws_logs.append([
                    log.id,
                    habit.id,
                    habit.name,
                    user.id,
                    user.name or user.username or "",
                    log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
                    status_text,
                    log.day_cycle or "",
                    to_msk(log.completed_at),
                ])
    
    # Auto-size columns for both sheets
    for ws in [ws_habits, ws_logs]:
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"mewego_habits_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    await update.message.reply_document(
        document=output,
        filename=filename,
        caption=f"📊 Экспорт: {total_habits} привычек, {total_logs} логов"
    )


async def notify_admin_new_user(bot, user: User) -> None:
    """Send notification to all admins about new user."""
    if not ADMIN_USERNAMES:
        return
    
    try:
        # Find all admins by username
        async with async_session() as session:
            result = await session.execute(
                select(User).where(User.username.in_(ADMIN_USERNAMES))
            )
            admins = result.scalars().all()
            
            if not admins:
                return
        
        habit_display = user.custom_habit if user.current_habit == "custom" else dict(HABITS).get(user.current_habit, user.current_habit)
        
        text = (
            "🆕 <b>Новый пользователь!</b>\n\n"
            f"👤 Имя: {user.name}\n"
            f"🎂 Возраст: {user.age}\n"
            f"📍 Город: {user.city}\n"
            f"💪 Активность: {user.activity_level}\n"
            f"🎯 Цель: {user.goal}\n"
            f"🏋️ Формат тренировок: {user.training_preference or '-'}\n"
            f"🏃 Привычка: {habit_display}\n"
            f"⏰ Напоминания: {user.reminder_time}\n"
            f"📊 Самоопознавание: {user.self_identification}\n"
        )
        
        # Send to all admins
        for admin in admins:
            try:
                await bot.send_message(
                    chat_id=admin.telegram_id,
                    text=text,
                    parse_mode="HTML"
                )
            except Exception:
                pass
    except Exception:
        pass  # Ignore errors


# =============================================================================
# BROADCAST FUNCTIONS
# =============================================================================

async def start_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start broadcast - ask admin for message to send."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return ConversationHandler.END
    
    cancel_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ Отмена", callback_data="admin:broadcast_cancel")]
    ])
    
    await query.message.reply_text(
        "📢 <b>Рассылка сообщения</b>\n\n"
        "Отправь мне сообщение, которое нужно разослать всем пользователям.\n\n"
        "Можешь отправить:\n"
        "• Текст\n"
        "• Фото с подписью\n"
        "• Видео с подписью\n\n"
        "<i>Для отмены нажми кнопку ниже</i>",
        parse_mode="HTML",
        reply_markup=cancel_keyboard
    )
    
    return BROADCAST_WAITING_MESSAGE


async def receive_broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Receive message for broadcast and ask for confirmation."""
    if not is_admin(update.effective_user.username):
        return ConversationHandler.END
    
    # Store message data in context
    context.user_data['broadcast_message'] = update.message
    
    # Count users who will receive the message
    async with async_session() as session:
        result = await session.execute(
            select(func.count(User.id)).where(User.onboarding_completed == True)
        )
        user_count = result.scalar()
    
    confirm_keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Да, отправить", callback_data="admin:broadcast_confirm"),
            InlineKeyboardButton("❌ Отмена", callback_data="admin:broadcast_cancel")
        ]
    ])
    
    await update.message.reply_text(
        f"📢 <b>Подтверждение рассылки</b>\n\n"
        f"Сообщение будет отправлено <b>{user_count}</b> пользователям.\n\n"
        f"<i>Подтверди отправку:</i>",
        parse_mode="HTML",
        reply_markup=confirm_keyboard
    )
    
    return BROADCAST_CONFIRM


async def confirm_broadcast_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle broadcast confirmation or cancellation."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        return ConversationHandler.END
    
    action = query.data.split(":")[-1]
    
    if action == "broadcast_cancel":
        context.user_data.pop('broadcast_message', None)
        await query.message.reply_text(
            "❌ Рассылка отменена",
            reply_markup=get_admin_panel_keyboard()
        )
        return ConversationHandler.END
    
    elif action == "broadcast_confirm":
        return await execute_broadcast(update, context)
    
    return ConversationHandler.END


async def execute_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Execute the broadcast to all users."""
    query = update.callback_query
    
    broadcast_message = context.user_data.get('broadcast_message')
    if not broadcast_message:
        await query.message.reply_text(
            "❌ Сообщение для рассылки не найдено",
            reply_markup=get_admin_panel_keyboard()
        )
        return ConversationHandler.END
    
    # Get all users who completed onboarding
    async with async_session() as session:
        result = await session.execute(
            select(User).where(User.onboarding_completed == True)
        )
        users = result.scalars().all()
    
    if not users:
        await query.message.reply_text(
            "❌ Нет пользователей для рассылки",
            reply_markup=get_admin_panel_keyboard()
        )
        return ConversationHandler.END
    
    # Send status message
    status_msg = await query.message.reply_text(
        f"📤 Начинаю рассылку... 0/{len(users)}"
    )
    
    success_count = 0
    fail_count = 0
    
    for i, user in enumerate(users):
        try:
            # Forward the message based on type
            if broadcast_message.photo:
                await context.bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=broadcast_message.photo[-1].file_id,
                    caption=broadcast_message.caption,
                    caption_entities=broadcast_message.caption_entities
                )
            elif broadcast_message.video:
                await context.bot.send_video(
                    chat_id=user.telegram_id,
                    video=broadcast_message.video.file_id,
                    caption=broadcast_message.caption,
                    caption_entities=broadcast_message.caption_entities
                )
            elif broadcast_message.document:
                await context.bot.send_document(
                    chat_id=user.telegram_id,
                    document=broadcast_message.document.file_id,
                    caption=broadcast_message.caption,
                    caption_entities=broadcast_message.caption_entities
                )
            elif broadcast_message.text:
                await context.bot.send_message(
                    chat_id=user.telegram_id,
                    text=broadcast_message.text,
                    entities=broadcast_message.entities
                )
            success_count += 1
        except Exception:
            fail_count += 1
        
        # Update status every 10 users
        if (i + 1) % 10 == 0:
            try:
                await status_msg.edit_text(
                    f"📤 Рассылка... {i + 1}/{len(users)}\n"
                    f"✅ Успешно: {success_count} | ❌ Ошибки: {fail_count}"
                )
            except Exception:
                pass
        
        # Small delay to avoid flood limits
        await asyncio.sleep(0.05)
    
    # Final status
    await status_msg.edit_text(
        f"✅ <b>Рассылка завершена!</b>\n\n"
        f"📊 Всего: {len(users)}\n"
        f"✅ Успешно: {success_count}\n"
        f"❌ Ошибки: {fail_count}",
        parse_mode="HTML"
    )
    
    await query.message.reply_text(
        "🔐 <b>Админ-панель</b>",
        parse_mode="HTML",
        reply_markup=get_admin_panel_keyboard()
    )
    
    # Clear stored message
    context.user_data.pop('broadcast_message', None)
    
    return ConversationHandler.END


async def cancel_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel broadcast via command."""
    context.user_data.pop('broadcast_message', None)
    await update.message.reply_text(
        "❌ Рассылка отменена",
        reply_markup=get_admin_panel_keyboard()
    )
    return ConversationHandler.END


# =============================================================================
# POLL FUNCTIONS
# =============================================================================

async def start_poll(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start poll creation - ask for question."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return ConversationHandler.END
    
    cancel_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ Отмена", callback_data="admin:poll_cancel")]
    ])
    
    await query.message.reply_text(
        "🗳 <b>Создание опроса</b>\n\n"
        "Шаг 1/2: Введи вопрос для опроса:",
        parse_mode="HTML",
        reply_markup=cancel_keyboard
    )
    
    return POLL_WAITING_QUESTION


async def receive_poll_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Receive poll question, ask for options."""
    if not is_admin(update.effective_user.username):
        return ConversationHandler.END
    
    question = update.message.text.strip()
    context.user_data['poll_question'] = question
    
    cancel_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ Отмена", callback_data="admin:poll_cancel")]
    ])
    
    await update.message.reply_text(
        "🗳 <b>Создание опроса</b>\n\n"
        f"Вопрос: <i>{question}</i>\n\n"
        "Шаг 2/2: Введи варианты ответов.\n"
        "Каждый вариант на новой строке:\n\n"
        "<code>Вариант 1\n"
        "Вариант 2\n"
        "Вариант 3</code>",
        parse_mode="HTML",
        reply_markup=cancel_keyboard
    )
    
    return POLL_WAITING_OPTIONS


async def receive_poll_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Receive poll options, ask for confirmation."""
    if not is_admin(update.effective_user.username):
        return ConversationHandler.END
    
    text = update.message.text.strip()
    options = [opt.strip() for opt in text.split('\n') if opt.strip()]
    
    if len(options) < 2:
        await update.message.reply_text(
            "❌ Нужно минимум 2 варианта ответа. Попробуй ещё раз:"
        )
        return POLL_WAITING_OPTIONS
    
    if len(options) > 10:
        await update.message.reply_text(
            "❌ Максимум 10 вариантов. Попробуй ещё раз:"
        )
        return POLL_WAITING_OPTIONS
    
    context.user_data['poll_options'] = options
    
    # Count users for preview
    async with async_session() as session:
        result = await session.execute(
            select(func.count(User.id)).where(User.onboarding_completed == True)
        )
        user_count = result.scalar()
    
    # Show preview
    question = context.user_data['poll_question']
    options_text = "\n".join([f"  {i+1}. {opt}" for i, opt in enumerate(options)])
    
    confirm_keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Отправить", callback_data="admin:poll_confirm"),
            InlineKeyboardButton("❌ Отмена", callback_data="admin:poll_cancel")
        ]
    ])
    
    await update.message.reply_text(
        f"🗳 <b>Предпросмотр опроса</b>\n\n"
        f"<b>{question}</b>\n\n"
        f"{options_text}\n\n"
        f"Будет отправлен <b>{user_count}</b> пользователям.\n\n"
        f"<i>Подтверди отправку:</i>",
        parse_mode="HTML",
        reply_markup=confirm_keyboard
    )
    
    return POLL_CONFIRM


async def confirm_poll_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle poll confirmation or cancellation."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        return ConversationHandler.END
    
    action = query.data.split(":")[-1]
    
    if action == "poll_cancel":
        context.user_data.pop('poll_question', None)
        context.user_data.pop('poll_options', None)
        await query.message.reply_text(
            "❌ Создание опроса отменено",
            reply_markup=get_admin_panel_keyboard()
        )
        return ConversationHandler.END
    
    elif action == "poll_confirm":
        return await execute_poll(update, context)
    
    return ConversationHandler.END


async def execute_poll(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Create poll in DB and send to all users."""
    query = update.callback_query
    
    question = context.user_data.get('poll_question')
    options = context.user_data.get('poll_options')
    
    if not question or not options:
        await query.message.reply_text(
            "❌ Ошибка: данные опроса не найдены",
            reply_markup=get_admin_panel_keyboard()
        )
        return ConversationHandler.END
    
    # Create poll in database
    async with async_session() as session:
        poll = Poll(
            question=question,
            options=json.dumps(options, ensure_ascii=False),
            created_by=query.from_user.id,
            is_active=True
        )
        session.add(poll)
        await session.commit()
        await session.refresh(poll)
        poll_id = poll.id
    
    # Get all users
    async with async_session() as session:
        result = await session.execute(
            select(User).where(User.onboarding_completed == True)
        )
        users = result.scalars().all()
    
    if not users:
        await query.message.reply_text(
            "❌ Нет пользователей для отправки",
            reply_markup=get_admin_panel_keyboard()
        )
        return ConversationHandler.END
    
    # Build voting keyboard
    vote_buttons = []
    for i, opt in enumerate(options):
        vote_buttons.append([
            InlineKeyboardButton(opt, callback_data=f"vote:{poll_id}:{i}")
        ])
    vote_keyboard = InlineKeyboardMarkup(vote_buttons)
    
    poll_text = f"🗳 <b>Опрос</b>\n\n{question}"
    
    # Send status message
    status_msg = await query.message.reply_text(
        f"📤 Отправляю опрос... 0/{len(users)}"
    )
    
    success_count = 0
    fail_count = 0
    
    for i, user in enumerate(users):
        try:
            await context.bot.send_message(
                chat_id=user.telegram_id,
                text=poll_text,
                parse_mode="HTML",
                reply_markup=vote_keyboard
            )
            success_count += 1
        except Exception:
            fail_count += 1
        
        if (i + 1) % 10 == 0:
            try:
                await status_msg.edit_text(
                    f"📤 Отправляю опрос... {i + 1}/{len(users)}\n"
                    f"✅ Успешно: {success_count} | ❌ Ошибки: {fail_count}"
                )
            except Exception:
                pass
        
        await asyncio.sleep(0.05)
    
    # Final status
    await status_msg.edit_text(
        f"✅ <b>Опрос #{poll_id} отправлен!</b>\n\n"
        f"📊 Всего: {len(users)}\n"
        f"✅ Успешно: {success_count}\n"
        f"❌ Ошибки: {fail_count}",
        parse_mode="HTML"
    )
    
    await query.message.reply_text(
        "🔐 <b>Админ-панель</b>",
        parse_mode="HTML",
        reply_markup=get_admin_panel_keyboard()
    )
    
    # Clear data
    context.user_data.pop('poll_question', None)
    context.user_data.pop('poll_options', None)
    
    return ConversationHandler.END


async def cancel_poll(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel poll creation via command."""
    context.user_data.pop('poll_question', None)
    context.user_data.pop('poll_options', None)
    await update.message.reply_text(
        "❌ Создание опроса отменено",
        reply_markup=get_admin_panel_keyboard()
    )
    return ConversationHandler.END


async def show_polls_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show list of all polls."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return
    
    async with async_session() as session:
        result = await session.execute(
            select(Poll).order_by(Poll.created_at.desc()).limit(10)
        )
        polls = result.scalars().all()
    
    if not polls:
        await query.message.reply_text(
            "📊 Опросов пока нет",
            reply_markup=get_admin_panel_keyboard()
        )
        return
    
    text = "📊 <b>Последние опросы:</b>\n\n"
    buttons = []
    
    for poll in polls:
        status = "🟢" if poll.is_active else "🔴"
        votes_count = len(poll.votes) if poll.votes else 0
        text += f"{status} <b>#{poll.id}</b>: {poll.question[:40]}...\n"
        text += f"   Голосов: {votes_count}\n\n"
        buttons.append([
            InlineKeyboardButton(
                f"📊 #{poll.id} - Результаты", 
                callback_data=f"admin:poll_results:{poll.id}"
            )
        ])
    
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="admin:back")])
    
    await query.message.reply_text(
        text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(buttons)
    )


async def show_poll_results(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show detailed results of a poll."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return
    
    poll_id = int(query.data.split(":")[-1])
    
    async with async_session() as session:
        result = await session.execute(
            select(Poll).where(Poll.id == poll_id)
        )
        poll = result.scalar_one_or_none()
        
        if not poll:
            await query.message.reply_text("❌ Опрос не найден")
            return
        
        # Get votes
        result = await session.execute(
            select(PollVote).where(PollVote.poll_id == poll_id)
        )
        votes = result.scalars().all()
    
    options = poll.get_options_list()
    
    # Count votes per option
    vote_counts = {i: 0 for i in range(len(options))}
    voters_per_option = {i: [] for i in range(len(options))}
    
    for vote in votes:
        if vote.option_index in vote_counts:
            vote_counts[vote.option_index] += 1
            voters_per_option[vote.option_index].append(vote.user_id)
    
    total_votes = sum(vote_counts.values())
    
    # Build results text
    status = "🟢 Активен" if poll.is_active else "🔴 Закрыт"
    text = f"📊 <b>Результаты опроса #{poll.id}</b>\n"
    text += f"Статус: {status}\n\n"
    text += f"<b>{poll.question}</b>\n\n"
    
    for i, option in enumerate(options):
        count = vote_counts[i]
        percent = (count / total_votes * 100) if total_votes > 0 else 0
        bar_length = int(percent / 10)
        bar = "█" * bar_length + "░" * (10 - bar_length)
        text += f"{i+1}. {option}\n"
        text += f"   {bar} {count} ({percent:.1f}%)\n\n"
    
    text += f"<b>Всего голосов: {total_votes}</b>"
    
    # Action buttons
    buttons = []
    if poll.is_active:
        buttons.append([
            InlineKeyboardButton("🔴 Закрыть опрос", callback_data=f"admin:close_poll:{poll_id}")
        ])
    buttons.append([
        InlineKeyboardButton("📋 Экспорт голосов", callback_data=f"admin:export_poll:{poll_id}")
    ])
    buttons.append([InlineKeyboardButton("⬅️ К списку опросов", callback_data="admin:polls_list")])
    
    await query.message.reply_text(
        text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(buttons)
    )


async def close_poll(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Close a poll to stop voting."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return
    
    poll_id = int(query.data.split(":")[-1])
    
    async with async_session() as session:
        result = await session.execute(
            select(Poll).where(Poll.id == poll_id)
        )
        poll = result.scalar_one_or_none()
        
        if poll:
            poll.is_active = False
            poll.closed_at = datetime.utcnow()
            await session.commit()
    
    await query.message.reply_text(
        f"✅ Опрос #{poll_id} закрыт",
        reply_markup=get_admin_panel_keyboard()
    )


async def export_poll_votes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Export poll votes to Excel."""
    query = update.callback_query
    await query.answer()
    
    if not is_admin(query.from_user.username):
        await query.answer("❌ Нет доступа", show_alert=True)
        return
    
    poll_id = int(query.data.split(":")[-1])
    
    async with async_session() as session:
        result = await session.execute(
            select(Poll).where(Poll.id == poll_id)
        )
        poll = result.scalar_one_or_none()
        
        if not poll:
            await query.message.reply_text("❌ Опрос не найден")
            return
        
        # Get votes with user info
        result = await session.execute(
            select(PollVote, User)
            .join(User, PollVote.user_id == User.telegram_id)
            .where(PollVote.poll_id == poll_id)
        )
        vote_data = result.all()
    
    options = poll.get_options_list()
    
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Опрос {poll_id}"
    
    ws.append(["Telegram ID", "Username", "Имя", "Ответ", "Дата голосования"])
    
    for vote, user in vote_data:
        option_text = options[vote.option_index] if vote.option_index < len(options) else "?"
        ws.append([
            vote.user_id,
            user.username or "",
            user.name or "",
            option_text,
            vote.voted_at.strftime("%Y-%m-%d %H:%M") if vote.voted_at else ""
        ])
    
    # Auto-size columns
    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"poll_{poll_id}_votes.xlsx"
    
    await query.message.reply_document(
        document=output,
        filename=filename,
        caption=f"📊 Экспорт голосов опроса #{poll_id}"
    )


async def vote_poll_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle user vote on a poll."""
    query = update.callback_query
    
    # Parse: vote:poll_id:option_index
    parts = query.data.split(":")
    if len(parts) != 3:
        await query.answer("❌ Ошибка")
        return
    
    poll_id = int(parts[1])
    option_index = int(parts[2])
    user_id = query.from_user.id
    
    async with async_session() as session:
        # Check if poll exists and is active
        result = await session.execute(
            select(Poll).where(Poll.id == poll_id)
        )
        poll = result.scalar_one_or_none()
        
        if not poll:
            await query.answer("❌ Опрос не найден", show_alert=True)
            return
        
        if not poll.is_active:
            await query.answer("❌ Опрос уже закрыт", show_alert=True)
            return
        
        # Check if already voted
        result = await session.execute(
            select(PollVote).where(
                PollVote.poll_id == poll_id,
                PollVote.user_id == user_id
            )
        )
        existing_vote = result.scalar_one_or_none()
        
        if existing_vote:
            await query.answer("⚠️ Ты уже голосовал в этом опросе!", show_alert=True)
            return
        
        # Record vote
        vote = PollVote(
            poll_id=poll_id,
            user_id=user_id,
            option_index=option_index
        )
        session.add(vote)
        await session.commit()
    
    options = poll.get_options_list()
    selected_option = options[option_index] if option_index < len(options) else "?"
    
    await query.answer(f"✅ Голос принят: {selected_option}", show_alert=True)
    
    # Update message to show user voted
    try:
        await query.edit_message_text(
            f"🗳 <b>Опрос</b>\n\n{poll.question}\n\n"
            f"✅ <i>Ты проголосовал: {selected_option}</i>",
            parse_mode="HTML"
        )
    except Exception:
        pass


async def admin_back_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle back button to admin panel."""
    query = update.callback_query
    await query.answer()
    
    await query.message.reply_text(
        "🔐 <b>Админ-панель</b>",
        parse_mode="HTML",
        reply_markup=get_admin_panel_keyboard()
    )

