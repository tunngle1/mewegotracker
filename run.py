#!/usr/bin/env python3
"""Entry point for bothost.ru and similar platforms."""
import sys
import os
import asyncio

# Add project root to Python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


async def run_import_once():
    """
    Import users from Excel file if it exists.
    Safely skips already existing users.
    """
    from pathlib import Path
    from datetime import datetime
    from zoneinfo import ZoneInfo
    from openpyxl import load_workbook
    from sqlalchemy import select
    from bot.database import async_session, init_db
    from bot.models import User
    
    # Check if Excel file exists
    excel_path = Path("DATABASE OLD/mewego_users_20260115_1122.xlsx")
    if not excel_path.exists():
        print("No Excel file to import, skipping...")
        return
    
    print("=" * 50)
    print("RUNNING USER IMPORT FROM EXCEL")
    print("=" * 50)
    
    MSK = ZoneInfo("Europe/Moscow")
    
    def parse_datetime(dt_str):
        if not dt_str or dt_str == "":
            return None
        try:
            msk_dt = datetime.strptime(str(dt_str), "%Y-%m-%d %H:%M")
            msk_dt = msk_dt.replace(tzinfo=MSK)
            return msk_dt.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
        except Exception:
            return None
    
    def safe_int(value, default=None):
        if value is None or value == "" or value == "None":
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    # Initialize database
    await init_db()
    
    # Load Excel
    wb = load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers)}
    print(f"Found columns: {list(col_map.keys())}")
    
    imported = 0
    skipped = 0
    
    async with async_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                telegram_id_val = row[col_map.get("Telegram ID", 1)]
                if not telegram_id_val:
                    continue
                
                telegram_id = int(telegram_id_val)
                
                result = await session.execute(
                    select(User).where(User.telegram_id == telegram_id)
                )
                if result.scalar_one_or_none():
                    skipped += 1
                    continue
                
                def get_col(name, default=None):
                    idx = col_map.get(name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return val if val and val != "None" else default
                    return default
                
                user = User(
                    telegram_id=telegram_id,
                    username=get_col("Username"),
                    name=get_col("Имя"),
                    age=safe_int(get_col("Возраст")),
                    city=get_col("Город"),
                    activity_level=get_col("Активность"),
                    goal=get_col("Цель"),
                    current_habit=get_col("Привычка"),
                    custom_habit=get_col("Своя привычка"),
                    day_cycle=safe_int(get_col("День цикла"), 1),
                    reminder_time=get_col("Время напоминания"),
                    onboarding_completed=(get_col("Онбординг") == "Да"),
                    self_identification=get_col("Самоопознавание"),
                    created_at=parse_datetime(get_col("Дата регистрации (МСК)")),
                    last_check_in=parse_datetime(get_col("Последняя отметка (МСК)")),
                )
                
                session.add(user)
                imported += 1
                print(f"Imported user: {user.name} ({telegram_id})")
                
            except Exception as e:
                print(f"Error: {e}")
                continue
        
        await session.commit()
    
    print(f"\n=== Users Import Complete ===")
    print(f"Imported: {imported}")
    print(f"Skipped (already exist): {skipped}")
    print("=" * 50)


async def run_habits_import():
    """
    Import habits from Excel file if it exists.
    Safely skips already existing habits.
    """
    from pathlib import Path
    from datetime import datetime
    from zoneinfo import ZoneInfo
    from openpyxl import load_workbook
    from sqlalchemy import select
    from bot.database import async_session, init_db
    from bot.models import User, Habit, ScheduleType
    
    # Check if habits Excel file exists
    excel_path = Path("DATABASE OLD/mewego_habits_20260115_1129.xlsx")
    if not excel_path.exists():
        print("No habits Excel file to import, skipping...")
        return
    
    print("=" * 50)
    print("RUNNING HABITS IMPORT FROM EXCEL")
    print("=" * 50)
    
    MSK = ZoneInfo("Europe/Moscow")
    
    def parse_datetime(dt_str):
        if not dt_str or dt_str == "":
            return None
        try:
            msk_dt = datetime.strptime(str(dt_str), "%Y-%m-%d %H:%M")
            msk_dt = msk_dt.replace(tzinfo=MSK)
            return msk_dt.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
        except Exception:
            return None
    
    def safe_int(value, default=None):
        if value is None or value == "" or value == "None":
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    await init_db()
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers)}
    print(f"Found columns: {list(col_map.keys())}")
    
    def get_col(row, name, default=None):
        idx = col_map.get(name)
        if idx is not None and idx < len(row):
            val = row[idx]
            return val if val and val != "None" else default
        return default
    
    imported = 0
    skipped = 0
    
    async with async_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                telegram_id = safe_int(get_col(row, "Telegram ID"))
                if not telegram_id:
                    continue
                
                result = await session.execute(
                    select(User).where(User.telegram_id == telegram_id)
                )
                user = result.scalar_one_or_none()
                
                if not user:
                    print(f"User not found: {telegram_id}")
                    continue
                
                habit_name = get_col(row, "Название привычки")
                if not habit_name:
                    continue
                
                # Check if habit exists
                result = await session.execute(
                    select(Habit).where(
                        Habit.user_id == user.id,
                        Habit.name == habit_name
                    )
                )
                if result.scalar_one_or_none():
                    skipped += 1
                    continue
                
                type_str = get_col(row, "Тип", "Ежедневно")
                schedule_type = ScheduleType.DAILY if "Ежедневно" in str(type_str) else ScheduleType.WEEKLY
                
                habit = Habit(
                    user_id=user.id,
                    name=habit_name,
                    schedule_type=schedule_type,
                    weekly_target=safe_int(get_col(row, "Цель (раз/нед)"), 7),
                    is_active=(get_col(row, "Активна") == "Да"),
                    created_at=parse_datetime(get_col(row, "Дата создания (МСК)")),
                )
                
                session.add(habit)
                imported += 1
                print(f"Imported habit: {habit_name} for {user.name}")
                
            except Exception as e:
                print(f"Error: {e}")
                continue
        
        await session.commit()
    
    print(f"\n=== Habits Import Complete ===")
    print(f"Imported: {imported}")
    print(f"Skipped (already exist): {skipped}")
    print("=" * 50)


# Now import and run the bot
from bot.main import main

if __name__ == "__main__":
    # Run imports first (safe to run multiple times - skips existing)
    asyncio.run(run_import_once())
    asyncio.run(run_habits_import())
    # Then start bot
    main()
